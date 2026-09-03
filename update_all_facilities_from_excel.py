import sys
import os
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

import openpyxl
import json
import requests
import concurrent.futures
from crypto_utils import encrypt_data

SUPABASE_URL = 'https://vijiacxcmtfekbmegjlf.supabase.co'
SECRET_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InZpamlhY3hjbXRmZWtibWVnamxmIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4NTgyMzgyNiwiZXhwIjoyMTAxMzk5ODI2fQ.Noa3eCRZLGLp67fRYu4ZlsFC4_d2X1C7KxQ_g2_zP00'
HEADERS = {
    'apikey': SECRET_KEY,
    'Authorization': f'Bearer {SECRET_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=representation'
}

def to_int(val, default=0):
    if val is None or val == '' or str(val).strip() == '':
        return default
    try:
        return int(float(str(val).strip()))
    except:
        return default

def to_str(val):
    if val is None or val == '' or str(val).strip() == 'None' or str(val).strip() == 'nan':
        return None
    s = str(val).strip()
    return s if s else None

def to_date_str(val):
    if val is None or val == '':
        return None
    s = str(val).strip()
    if len(s) >= 10 and s[:4].isdigit():
        return s[:10]
    return None

print("[1/4] DB.xlsx 파일 로딩 중...")
wb = openpyxl.load_workbook('DB.xlsx', data_only=True)
sheet = wb['통합'] if '통합' in wb.sheetnames else wb.worksheets[0]

# 현재 DB 컬럼 목록 조회
res_cols = requests.get(f'{SUPABASE_URL}/rest/v1/facilities?limit=1', headers=HEADERS)
db_existing_cols = set(res_cols.json()[0].keys()) if res_cols.status_code == 200 and res_cols.json() else set()
print(f"현재 DB 컬럼 목록 ({len(db_existing_cols)}개): {db_existing_cols}")

facilities_cache_list = []
db_update_tasks = []

print("[2/4] 엑셀 데이터 파싱 및 암호화 처리 중...")
for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
    if row_idx == 0:
        continue  # 헤더 건너뛰기
    
    key = to_str(row[72])  # BU열 (KEY)
    if not key or not key.startswith('K'):
        continue

    fac_name = to_str(row[4])
    fac_cat = to_str(row[1])
    comp_status = to_str(row[2])
    inv_status = to_str(row[3])
    addr_jibun = to_str(row[5])
    addr_doro = to_str(row[6])
    dong = to_str(row[7])
    bldg_num = to_str(row[8])
    perm_date = to_date_str(row[10])
    appr_date = to_date_str(row[11])
    new_old = to_str(row[12])
    own_type = to_str(row[13])

    # 주차구역 P~U
    p_req = to_int(row[15])   # P: 의무_면수
    p_inst = to_int(row[16])  # Q: 면수합
    p_grd = to_int(row[17])   # R: 지상면수
    p_ugrd = to_int(row[18])  # S: 지하면수
    p_uninst = to_int(row[19])# T: 미설치면수
    p_stat = to_str(row[20])  # U: 주차면수 이행여부

    # 충전시설 V~AB
    c_req = to_int(row[21])   # V: 의무_시설
    c_fast_req = to_int(row[22]) # W: 의무_급속
    c_inst = to_int(row[23])  # X: 시설합
    c_fast = to_int(row[24])  # Y: 급속기수
    c_slow = to_int(row[25])  # Z: 완속기수
    c_uninst = to_int(row[26])# AA: 미설치기수
    c_stat = to_str(row[27])  # AB: 충전시설 이행여부

    # 관리자 정보 (규칙: 암호화 필수)
    mgr_name = to_str(row[35])
    mgr_contact = to_str(row[36]) or to_str(row[37])
    mgr_name_enc = encrypt_data(mgr_name) if mgr_name else None
    mgr_contact_enc = encrypt_data(mgr_contact) if mgr_contact else None
    mgmt_body = to_str(row[69])

    # 추가 엑셀 필드
    covenant = to_str(row[9])
    general_mixed = to_str(row[14])
    total_hh = to_int(row[28])
    ev_reg = to_int(row[29])
    c_reported = to_str(row[30])
    ins_enrolled = to_str(row[31])
    parallel_stat = to_str(row[32])
    parallel_cnt = to_int(row[33])
    fire_manual = to_str(row[34])
    final_conclusion = to_str(row[68])
    job_parking_cnt = to_int(row[70])
    order_addr = to_str(row[71])

    # 조사 차수 상세
    surveys = {
        '1st': {'method': to_str(row[38]), 'date': to_date_str(row[39]), 'investigator': to_str(row[40]), 'check': to_str(row[41]), 'plan': to_str(row[42]), 'note': to_str(row[43])},
        '2nd': {'method': to_str(row[44]), 'date': to_date_str(row[45]), 'investigator': to_str(row[46]), 'check': to_str(row[47]), 'plan': to_str(row[48]), 'note': to_str(row[49])},
        '3rd': {'method': to_str(row[50]), 'date': to_date_str(row[51]), 'investigator': to_str(row[52]), 'check': to_str(row[53]), 'plan': to_str(row[54]), 'note': to_str(row[55])},
        '4th': {'method': to_str(row[56]), 'date': to_date_str(row[57]), 'investigator': to_str(row[58]), 'check': to_str(row[59]), 'plan': to_str(row[60]), 'note': to_str(row[61])},
        '5th': {'method': to_str(row[62]), 'date': to_date_str(row[63]), 'investigator': to_str(row[64]), 'check': to_str(row[65]), 'plan': to_str(row[66]), 'note': to_str(row[67])},
    }

    # DB 전송용 페이로드 (현재 DB에 존재하는 컬럼만 포함)
    full_db_candidate = {
        'facility_key': key,
        'facility_name': fac_name,
        'facility_category': fac_cat,
        'compliance_status': comp_status,
        'investigation_status': inv_status,
        'address_jibun': addr_jibun,
        'address_doro': addr_doro,
        'dong_name': dong,
        'building_register_num': bldg_num,
        'permission_date': perm_date,
        'approval_date': appr_date,
        'is_new_building': new_old,
        'facility_ownership_type': own_type,
        'parking_required_cnt': p_req,
        'parking_installed_cnt': p_inst,
        'parking_ground_cnt': p_grd,
        'parking_underground_cnt': p_ugrd,
        'parking_uninstalled_cnt': p_uninst,
        'parking_status': p_stat,
        'charger_required_cnt': c_req,
        'charger_fast_req_cnt': c_fast_req,
        'charger_installed_cnt': c_inst,
        'charger_fast_cnt': c_fast,
        'charger_slow_cnt': c_slow,
        'charger_uninstalled_cnt': c_uninst,
        'charger_status': c_stat,
        'manager_name_encrypted': mgr_name_enc,
        'manager_contact_encrypted': mgr_contact_enc,
        'management_body': mgmt_body,
        'total_households': total_hh,
        'ev_registered_cnt': ev_reg,
        'charger_reported': c_reported,
        'insurance_enrolled': ins_enrolled,
        'parallel_parking_status': parallel_stat,
        'parallel_parking_cnt': parallel_cnt,
        'fire_manual_distributed': fire_manual,
    }

    # 실제 DB 컬럼에 맞는 것만 필터링
    db_payload = {k: v for k, v in full_db_candidate.items() if k in db_existing_cols}
    db_update_tasks.append((key, db_payload))

    # 프론트엔드 캐시용 완전 객체
    cache_item = {
        **full_db_candidate,
        'manager_name_decrypted': mgr_name or '',
        'manager_contact_decrypted': mgr_contact or '',
        'manager_name': mgr_name or '',
        'manager_contact': mgr_contact or '',
        'covenant': covenant,
        'general_mixed': general_mixed,
        'final_conclusion': final_conclusion,
        'job_parking_cnt': job_parking_cnt,
        'order_addr': order_addr,
        'surveys': surveys
    }
    facilities_cache_list.append(cache_item)

print(f"파싱 완료: 총 {len(facilities_cache_list)}개 시설 데이터 준비됨.")

# [3/4] Supabase DB 일괄 업데이트
print(f"[3/4] Supabase DB에 479개 시설 일괄 동기화(PATCH) 전송 중...")
def patch_facility(task):
    key, payload = task
    try:
        res = requests.patch(f"{SUPABASE_URL}/rest/v1/facilities?facility_key=eq.{key}", headers=HEADERS, json=payload, timeout=10)
        return res.status_code in [200, 204]
    except:
        return False

success_count = 0
fail_count = 0
with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
    results = list(executor.map(patch_facility, db_update_tasks))
    success_count = sum(1 for r in results if r)
    fail_count = len(results) - success_count

print(f"DB 업데이트 결과: 성공 {success_count}건, 실패 {fail_count}건")

# [4/4] 로컬 캐시 facilities_cache.json 업데이트
print("[4/4] 로컬 캐시 facilities_cache.json 업데이트 중...")
with open('facilities_cache.json', 'w', encoding='utf-8') as f:
    json.dump(facilities_cache_list, f, ensure_ascii=False, indent=2)
print("facilities_cache.json 저장 완료!")

print("\n=== 작업 완료! ===")
